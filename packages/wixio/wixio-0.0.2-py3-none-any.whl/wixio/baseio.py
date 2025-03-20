# -*- encoding: utf-8 -*-
"""
@File    :   baseio.py
@Time    :   2025/2/25 下午8:21
@Author  :   Li Jiawei
@Version :   1.0
@Contact :   Li.J.W.adrian421@hotmail.com
@License :   (C)Copyright 2023-2030
@Desc    :   None
@Brief   :
"""

import os, sys, argparse
from tqdm import tqdm
from itertools import groupby, tee
from datetime import timedelta
from collections import Counter
import codecs, json, \
        time, datetime, pickle, \
        re, random, numpy as np, \
        pandas as pd, xml.dom.minidom as xmldom, string
import inspect
import platform
import yaml
import itertools
from pdb import set_trace as stop
from typing import Any, List, Tuple, Union


class wixio(object):
        def __init__(self, data=None, len=None, doc=None, info=None, details=None):
                """
                Data structure for general data loaded from various files.

                Parameter:
                ----------
                r: Axis-angle rotation tensor of shape [batch_size, 1, 3].

                Return:
                -------
                Rotation matrix of shape [batch_size, 3, 3].
                """
                self.data = data
                self.len = len
                self.doc = doc
                self.info = info
                self.details = details

        @classmethod
        def tojson(cls, data, path, encoding='utf-8'):
                with codecs.open(path, 'w', encoding=encoding) as f:
                        json.dump(data, f, ensure_ascii=False)
                        print(f'Finished dumping json file at {path}!')

        @classmethod
        def totxt(cls, data: list, path, encoding='utf-8'):
                assert isinstance(data, list)
                with codecs.open(path, 'w', encoding=encoding) as f:
                        for item in data:
                                f.write(item + '\n')
                print(f'Finished dumping txt file at {path}!')

        @classmethod
        def tocsv(cls, data: dict, path, encoding='utf-8'):
                assert isinstance(data, dict)
                df = pd.DataFrame(data)
                df.to_csv(path, sep=',', header=True, index=False)
                print(f"Finished dumping csv file at {path}!")

        @classmethod
        def toxlsx(cls, data: dict, path, encoding='utf-8'):
                assert isinstance(data, dict)
                df = pd.DataFrame(data)
                df.to_excel(path, engine='xlsxwriter', index=False, header=True)
                print(f"Finished dumping xlsx file at {path}!")

        @classmethod
        def savepkl(cls, data, path):
                with open(path, 'wb') as pklfile:
                        pickle.dump(data, pklfile)
                print(f'Finished dumping pkl file at {path}!')

        @classmethod
        def read(cls, path, split_mark='\t', encoding='utf-8', sheet_id=0, dform=None, **kwargs):
                if path.endswith('.txt'):
                        new_lines = []
                        size = 0
                        with open(path, 'r', encoding=encoding, errors='ignore') as file:
                                lines = file.readlines()
                                size = len(lines)
                                for i in range(size):
                                        line = lines[i].strip().split(split_mark)
                                        if len(line) == 1: line = line[0]
                                        new_lines.append(line)
                        return cls(data=DataWrapper(new_lines), len=size, doc=path, info='txt')
                elif path.endswith('.json'):
                        with open(path, 'r', encoding=encoding) as f:
                                data = json.loads(f.read())
                                size = len(data)
                        return cls(data=DataWrapper(data), len=size, doc=path, info='json')
                elif path.endswith('.xml'):
                        domtree = xmldom.parse(path)
                        rootnode = domtree.documentElement
                        instruction = ("XML文件读取后,采用以下函数读取标签:\n"
                                       "对于一个为'paragraph'的标签: nodes_paragraph = rootnode.getElementsByTagName('paragraph')\n"
                                       "之后通过enumerate遍历里面的内容:\n"
                                       "for idx, node in enumerate(nodes_paragraph): ......\n")
                        return cls(data=DataWrapper(rootnode), len=0, doc=path, info=instruction)
                elif path.endswith('.csv'):
                        f = open(path, 'r', encoding=encoding, errors='ignore')
                        main_content = pd.DataFrame(pd.read_csv(f, header=0, low_memory=False))
                        main_content = main_content.where(main_content.notnull(), 'NAN')
                        info = ("注意:\n"
                                "读取的DataFrame经过预处理,将所有空白项填充了字符串 'NAN' \n"
                                "需要调用哪列,就 data.lines[列名]\n")
                        return cls(data=DataWrapper(main_content), len=len(main_content), doc=path, info=info)
                elif path.endswith('.xlsx'):
                        import openpyxl
                        # 读取xlsx文件
                        wb = openpyxl.load_workbook(path)
                        ws = wb.worksheets[sheet_id]
                        # sheet页属性：表名、最大行数、最大列数
                        info = 'titles: {}, max_row: {}, max_column: {}'.format(ws.title, ws.max_row, ws.max_column)
                        columns = ws.columns
                        all_data = []
                        for col in columns:
                                a_col = []
                                for item in col: a_col.append(item.value)
                                all_data.append(a_col)
                        headers = [col[0] for col in all_data]
                        return cls(data=DataWrapper(all_data), len=len(all_data[0]), doc=path, info=info, details={'header': headers})
                elif path.endswith('.pkl'):
                        f = open(path, 'rb')
                        content = pickle.load(f)
                        return cls(data=DataWrapper(content), len=len(content), doc=path, info='pkl')
                elif path.endswith(".yaml"):
                        f = open(path, 'r', encoding='utf-8')
                        cont = f.read()
                        fname = path.split('/')[-1].split('.')[0]
                        res = yaml.load(cont, yaml.FullLoader)
                        attri_len = len(res.keys())
                        all_attributes = '\t'.join([f"({idx}) {x}" for idx, x in enumerate(list(res.keys()))])
                        res = type(f"{fname}", (), res)  # meta class
                        if dform is None or dform == 'cls':
                                return cls(data=DataWrapper(res), len=attri_len, doc=path, info=f'yaml\n[Attributes]: {all_attributes}')
                        if dform == 'dict':
                                res_dict = {}
                                for k, v in res.__dict__.items():
                                        if not k.startswith("__"):
                                                res_dict[k] = res.__dict__.get(k, v)
                                return cls(data=DataWrapper(res_dict), len=attri_len, doc=path, info=f'yaml\n[Attributes]: {all_attributes}')

        @classmethod
        def rm_nan(cls, df):
                assert isinstance(df, pd.DataFrame), f"input is not pd.DataFrame"
                return df.where(df.notnull(), 'NAN')


class DataWrapper:
        """数据包装器，实现属性式访问的增强功能"""

        def __init__(self, data):
                self._data = data
                self._type = type(data)

                # 根据数据类型动态添加属性
                if isinstance(data, list):
                        self.len = len(data)
                elif isinstance(data, dict):
                        self.keys = list(data.keys())
                        self.kvs = list(data.items())
                elif hasattr(data, '__dict__'):  # 处理类实例
                        self.keys = list(data.__dict__.keys())
                        self.kvs = list(data.__dict__.items())

        @property
        def type(self):
                """获取数据类型"""
                return self._type

        def __getattr__(self, name):
                """转发未处理的属性到原始数据"""
                return getattr(self._data, name)

        def __getitem__(self, index):
                """保留索引访问能力"""
                return self._data[index]

        def __iter__(self):
                """保留迭代能力"""
                return iter(self._data)

        def __repr__(self):
                """保持原始数据表示形式"""
                return repr(self._data)


def frame_save(data:list, title:str, info:str, datasize=None):
    generate_date = datetime.date.today()
    generate_date.isoformat()
    generate_date = str(generate_date)
    if datasize is None:
        return {'title':title, 'date':generate_date, 'info':info, 'size':f"{len(data)}", 'content':data}
    else:
        return {'title':title, 'date':generate_date, 'info':info, 'size':f"{datasize}", 'content':data}

