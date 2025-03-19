import os
import json
import boto3
import openpyxl
from datetime import datetime
import logging

logger = logging.getLogger(__name__)

# Config file path (stored in the user's home directory)
CONFIG_PATH = os.path.expanduser("~/.aws_report_config.json")

def load_config():
    """Load AWS configuration from a JSON file, if it exists."""
    if os.path.exists(CONFIG_PATH):
        with open(CONFIG_PATH, "r") as file:
            return json.load(file)
    return {}

# Load configuration from file
config = load_config()

# Get AWS settings with priority: Function Parameter > Environment Variable > Config File > Default Value
AWS_REGION = os.getenv("AWS_REGION", config.get("AWS_REGION", "us-east-1"))
S3_BUCKET_NAME = os.getenv("S3_BUCKET_NAME", config.get("S3_BUCKET_NAME", "rk-test-reports"))
S3_FOLDER = os.getenv("S3_FOLDER", config.get("S3_FOLDER", "reports/"))

# Initialize S3 Client
s3_client = boto3.client("s3", region_name=AWS_REGION)

def fetch_aws_data(boto_session, aws_region=None):
    """
    Fetch AWS inventory data (EC2, S3, IAM Users) using a boto3 session.
    Allows specifying the AWS region dynamically.
    """
    aws_region = aws_region or AWS_REGION  # Use function parameter if provided, else fallback

    ec2_client = boto_session.client("ec2", region_name=aws_region)
    s3_client = boto_session.client("s3")
    iam_client = boto_session.client("iam")

    # Fetch EC2 instances
    ec2_instances = ec2_client.describe_instances()
    instance_data = [
        {
            "instance_id": i["InstanceId"],
            "instance_type": i["InstanceType"],
            "state": i["State"]["Name"],
            "region": aws_region
        }
        for r in ec2_instances["Reservations"]
        for i in r["Instances"]
    ]

    # Fetch S3 Buckets
    s3_buckets = [{"bucket_name": b["Name"], "creation_date": b["CreationDate"]} for b in s3_client.list_buckets()["Buckets"]]

    # Fetch IAM Users
    iam_users = [{"username": u["UserName"], "arn": u["Arn"], "create_date": u["CreateDate"]} for u in iam_client.list_users()["Users"]]

    return instance_data, s3_buckets, iam_users

def generate_excel_report(account_id, boto_session, aws_region=None, s3_bucket=None, s3_folder=None):
    """
    Generate an Excel report from AWS data with dynamic configuration options.
    """
    aws_region = aws_region or AWS_REGION
    s3_bucket = s3_bucket or S3_BUCKET_NAME
    s3_folder = s3_folder or S3_FOLDER

    ec2_instances, s3_buckets, iam_users = fetch_aws_data(boto_session, aws_region)

    wb = openpyxl.Workbook()

    # EC2 Sheet
    ws1 = wb.active
    ws1.title = "EC2 Instances"
    ws1.append(["Instance ID", "Instance Type", "State", "Region"])
    for instance in ec2_instances:
        ws1.append([instance["instance_id"], instance["instance_type"], instance["state"], instance["region"]])

    # S3 Sheet
    ws2 = wb.create_sheet(title="S3 Buckets")
    ws2.append(["Bucket Name", "Creation Date"])
    for bucket in s3_buckets:
        ws2.append([bucket["bucket_name"], bucket["creation_date"].strftime("%Y-%m-%d %H:%M:%S")])

    # IAM Sheet
    ws3 = wb.create_sheet(title="IAM Users")
    ws3.append(["Username", "ARN", "Creation Date"])
    for user in iam_users:
        ws3.append([user["username"], user["arn"], user["create_date"].strftime("%Y-%m-%d %H:%M:%S")])

    # Save file
    file_name = f"aws_inventory_{account_id}_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
    file_path = f"/tmp/{file_name}"
    wb.save(file_path)

    return file_path, file_name

def upload_to_s3(file_path, file_name, s3_bucket=None, s3_folder=None):
    """
    Uploads the Excel report to an S3 bucket with a pre-signed URL.
    """
    s3_bucket = s3_bucket or S3_BUCKET_NAME
    s3_folder = s3_folder or S3_FOLDER
    file_key = f"{s3_folder}{file_name}"

    # Upload file to S3
    s3_client.upload_file(file_path, s3_bucket, file_key)

    # Generate pre-signed URL (valid for 24 hours)
    presigned_url = s3_client.generate_presigned_url(
        "get_object",
        Params={"Bucket": s3_bucket, "Key": file_key},
        ExpiresIn=86400
    )

    return presigned_url