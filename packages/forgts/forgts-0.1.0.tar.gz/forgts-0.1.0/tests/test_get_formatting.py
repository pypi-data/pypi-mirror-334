import unittest
from unittest.mock import Mock, patch
import pandas as pd
from openpyxl.styles import Font, Fill, Border, Side, Color
from forgts.get_formatting import ExcelFormatter


class TestExcelFormatter(unittest.TestCase):

    def setUp(self):
        """Set up test fixtures before each test method."""
        self.filepath = "test.xlsx"
        self.formatter = ExcelFormatter(self.filepath)
        
    def test_init(self):
        """Test the initialization of ExcelFormatter."""
        formatter = ExcelFormatter(self.filepath)
        self.assertEqual(formatter.xlfilepath, self.filepath)
        self.assertEqual(formatter.sheet, 0)
        
        # Test with sheet name
        formatter = ExcelFormatter(self.filepath, "Sheet1")
        self.assertEqual(formatter.sheet, "Sheet1")
        
        # Test with sheet index
        formatter = ExcelFormatter(self.filepath, 2)
        self.assertEqual(formatter.sheet, 2)
    
    @patch('openpyxl.load_workbook')
    def test_load_workbook_with_index(self, mock_load):
        """Test loading workbook with sheet index."""
        mock_wb = Mock()
        mock_ws = Mock()
        mock_wb.worksheets = [mock_ws]
        mock_load.return_value = mock_wb
        
        formatter = ExcelFormatter(self.filepath, 0)
        formatter.load_workbook()
        
        mock_load.assert_called_once_with(self.filepath)
        self.assertEqual(formatter.workbook, mock_wb)
        self.assertEqual(formatter.worksheet, mock_ws)
    
    @patch('openpyxl.load_workbook')
    def test_load_workbook_with_name(self, mock_load):
        """Test loading workbook with sheet name."""
        mock_wb = Mock()
        mock_ws = Mock()

        # Explicitly set the __getitem__ method on the mock workbook
        mock_wb.__getitem__ = Mock(return_value=mock_ws)
        mock_load.return_value = mock_wb

        formatter = ExcelFormatter(self.filepath, "Sheet1")
        formatter.load_workbook()

        mock_load.assert_called_once_with(self.filepath)
        mock_wb.__getitem__.assert_called_once_with("Sheet1")
        self.assertEqual(formatter.workbook, mock_wb)
        self.assertEqual(formatter.worksheet, mock_ws)
    
    def test_get_dimensions_with_data(self):
        """Test getting dimensions when cells have values."""
        mock_cell1 = Mock()
        mock_cell1.value = "Data"
        mock_cell2 = Mock()
        mock_cell2.value = None
        mock_cell3 = Mock()
        mock_cell3.value = "More Data"
        
        # Create a 2x2 grid with data in cells (1,1) and (2,2)
        self.formatter.worksheet = Mock()
        self.formatter.worksheet.iter_rows.return_value = [
            [mock_cell1, mock_cell2],
            [mock_cell2, mock_cell3]
        ]
        
        self.formatter.get_dimensions()
        
        # The dimensions should be (2, 2) because the furthest non-empty cell is at (2, 2)
        self.assertEqual(self.formatter.dimensions, (2, 2))
    
    def test_get_dimensions_empty_sheet(self):
        """Test getting dimensions when the sheet is empty."""
        # All cells are empty
        mock_cell = Mock()
        mock_cell.value = None
        
        self.formatter.worksheet = Mock()
        self.formatter.worksheet.iter_rows.return_value = [
            [mock_cell, mock_cell],
            [mock_cell, mock_cell]
        ]
        self.formatter.worksheet.max_row = 2
        self.formatter.worksheet.max_column = 2
        
        self.formatter.get_dimensions()
        
        # Should fall back to worksheet dimensions
        self.assertEqual(self.formatter.dimensions, (2, 2))
    
    @patch('pandas.read_excel')
    def test_read_and_validate_spreadsheet_valid(self, mock_read_excel):
        """Test reading a valid spreadsheet."""
        mock_df = pd.DataFrame({
            'Col1': [1, 2, 3],
            'Col2': [4, 5, 6]
        })
        mock_read_excel.return_value = mock_df
        
        self.formatter.read_and_validate_spreadsheet()
        
        mock_read_excel.assert_called_once()
        pd.testing.assert_frame_equal(self.formatter.spreadsheet_data, mock_df)
    
    @patch('pandas.read_excel')
    def test_read_and_validate_spreadsheet_invalid(self, mock_read_excel):
        """Test reading a spreadsheet with invalid headers."""
        mock_df = pd.DataFrame({
            'Col1': [1, 2, 3],
            '...1': [4, 5, 6]  # Invalid column name
        })
        mock_read_excel.return_value = mock_df
        
        with self.assertRaises(ValueError):
            self.formatter.read_and_validate_spreadsheet()
    
    def test_extract_single_cell_formatting(self):
        """Test extraction of formatting from a single cell."""
        # Create a mock cell with specific formatting
        mock_cell = Mock()
        
        # Set up font properties
        mock_font = Mock()
        mock_font.bold = True
        mock_font.italic = False
        mock_font.underline = 'single'
        mock_font.strike = True
        mock_font.color = Mock()
        mock_font.color.type = "rgb"
        mock_font.color.rgb = "1"
        mock_cell.font = mock_font
        
        # Set up fill properties
        mock_fill = Mock()
        mock_fill.start_color = Mock()
        mock_fill.start_color.type = "rgb"
        mock_fill.start_color.rgb = "00FF9900"
        mock_cell.fill = mock_fill
        
        # Set up border properties
        mock_border = Mock()
        
        mock_top_side = Mock()
        mock_top_side.style = "thin"
        mock_top_side.color = Mock()
        mock_top_side.color.type = "rgb"
        mock_top_side.color.rgb = "0000FF"
        
        mock_right_side = None
        mock_bottom_side = None
        mock_left_side = None
        
        mock_border.top = mock_top_side
        mock_border.right = mock_right_side
        mock_border.bottom = mock_bottom_side
        mock_border.left = mock_left_side
        
        mock_cell.border = mock_border
        
        # Extract formatting
        row, col = 1, 1
        result = ExcelFormatter.extract_single_cell_formatting(mock_cell, row, col)
        
        # Assert that all properties were correctly extracted
        expected = {
            'row': 1,
            'col': 1,
            'bold': True,
            'italic': False,
            'underlined':'single',
            'strikethrough': True,
            'border_top_style': "thin",
            'border_right_style': None,
            'border_bottom_style': None,
            'border_left_style': None,
            'hl_color': '00FF9900',
            'text_clr': '1',
            'border_top_clr': '0000FF',
            'border_right_clr': None,
            'border_bottom_clr': None,
            'border_left_clr': None,
        }
        self.assertEqual(result, expected)
    
    def test_get_color(self):
        test_cases = [
            ("rgb", {"rgb": "FF0000"}, "FF0000"),
            ("indexed", {"indexed": 0}, "00000000"),
            ("theme", {"theme": 1}, "1"),
            ("auto", {}, "auto"),
            (None, {}, None),
        ]
        
        for color_type, attributes, expected in test_cases:
            with self.subTest(color_type=color_type):
                mock_color = Mock()
                mock_color.type = color_type
                for attr, value in attributes.items():
                    setattr(mock_color, attr, value)
                
                if color_type == "indexed":
                    with patch('openpyxl.styles.colors.COLOR_INDEX', {0: "00000000"}):
                        result = ExcelFormatter.get_color(mock_color)
                else:
                    result = ExcelFormatter.get_color(mock_color)
                
                self.assertEqual(result, expected)
    
    def test_create_empty_cell_record(self):
        """Test creation of empty cell record."""
        row, col = 4, 2
        result = ExcelFormatter.create_empty_cell_record(row, col)
        
        expected = {
            'row': 4,
            'col': 2,
            'bold': None,
            'italic': None,
            'underlined': None,
            'hl_color': None,
            'strikethrough': None,
            'text_clr': None,
            'border_top_style': None,
            'border_top_clr': None,
            'border_right_style': None,
            'border_right_clr': None,
            'border_bottom_style': None,
            'border_bottom_clr': None,
            'border_left_style': None,
            'border_left_clr': None,
        }
        self.assertEqual(result, expected)
    
    def test_complete_formatting_data(self):
        """Test completing formatting data by filling in missing entries."""
        # Create a formatting dataframe with some missing rows
        data = [
            {'row': 1, 'col': 1, 'bold': True, 'italic': False, 'underlined': None, 
             'hl_color': None, 'strikethrough': None, 'text_clr': None,
             'border_top_style': None, 'border_top_clr': None, 'border_right_style': None, 
             'border_right_clr': None, 'border_bottom_style': None, 'border_bottom_clr': None, 
             'border_left_style': None, 'border_left_clr': None},
            # Row 2, Col 1 is missing
            {'row': 3, 'col': 1, 'bold': False, 'italic': True, 'underlined': None, 
             'hl_color': None, 'strikethrough': None, 'text_clr': None,
             'border_top_style': None, 'border_top_clr': None, 'border_right_style': None, 
             'border_right_clr': None, 'border_bottom_style': None, 'border_bottom_clr': None, 
             'border_left_style': None, 'border_left_clr': None},
        ]
        self.formatter.formatting_data = pd.DataFrame(data)
        
        self.formatter.complete_formatting_data()
        
        # Check that all rows exist
        result_df = self.formatter.formatting_data
        self.assertEqual(len(result_df), 3)  # 3 rows for column 1
        
        # Check that the missing entry was filled
        missing_row2_col1 = result_df[(result_df['row'] == 2) & (result_df['col'] == 1)]
        self.assertEqual(len(missing_row2_col1), 1)
        self.assertIsNone(missing_row2_col1.iloc[0]['bold'])
    
    def test_process_single_column_formatting(self):
        """Test processing formatting for a single column."""
        # Setup
        self.formatter.spreadsheet_data = pd.DataFrame({
            'Col1': [1, 2],
            'Col2': [3, 4]
        })
        
         # Mock formatting data
        data = [
            {'row': 1, 
             'col': 1, 
             'bold': False, 
             'italic': False, 
             'underlined': None, 
             'hl_color': 'FF0000', 
             'strikethrough': None, 
             'text_clr': None,
             'border_top_style': None, 
             'border_top_clr': None, 
             'border_right_style': None, 
             'border_right_clr': None, 
             'border_bottom_style': None, 
             'border_bottom_clr': None, 
             'border_left_style': None, 
             'border_left_clr': None},
            {'row': 2, 
             'col': 1, 
             'bold': True, 
             'italic': False, 
             'underlined': None, 
             'hl_color': None, 
             'strikethrough': None, 
             'text_clr': 'FF0000',
             'border_top_style': None, 
             'border_top_clr': None, 
             'border_right_style': None, 
             'border_right_clr': None, 
             'border_bottom_style': None, 
             'border_bottom_clr': None, 
             'border_left_style': None, 
             'border_left_clr': None},
        ]
        
        self.formatter.formatting_data = pd.DataFrame(data)
        result = self.formatter.process_single_column_formatting('Col1')

        # Verify the result
        self.assertEqual(list(result['target_var'].unique()), ['Col1'])
        self.assertTrue(all(row in ['1', '2'] for row in result['rowid']))
    
    @patch('pandas.concat')
    def test_process_column_formatting(self, mock_concat):
        """Test processing formatting for all columns."""
        # Setup
        self.formatter.spreadsheet_data = pd.DataFrame({
            'Col1': [1, 2],
            'Col2': [3, 4]
        })
        
        col1_result = pd.DataFrame({'col1_data': [1, 2]})
        col2_result = pd.DataFrame({'col2_data': [3, 4]})
        
        with patch.object(self.formatter, 'process_single_column_formatting') as mock_process:
            mock_process.side_effect = [col1_result, col2_result]
            mock_concat.return_value = pd.DataFrame({'combined': [1, 2, 3, 4]})
            
            result = self.formatter.process_column_formatting()
            
            # Verify the method was called for each column
            self.assertEqual(mock_process.call_count, 2)
            mock_process.assert_any_call('Col1')
            mock_process.assert_any_call('Col2')
    
    @patch.object(ExcelFormatter, 'load_workbook')
    @patch.object(ExcelFormatter, 'get_dimensions')
    @patch.object(ExcelFormatter, 'read_and_validate_spreadsheet')
    @patch.object(ExcelFormatter, 'extract_cell_formatting')
    @patch.object(ExcelFormatter, 'complete_formatting_data')
    @patch.object(ExcelFormatter, 'process_column_formatting')
    def test_run(self, mock_process_column, mock_complete, mock_extract, 
                mock_read, mock_get_dim, mock_load):
        """Test running the entire process."""
        expected_result = pd.DataFrame({'result': [1, 2, 3]})
        mock_process_column.return_value = expected_result
        
        result = self.formatter.run()
        
        # Verify all methods were called in the correct order
        mock_load.assert_called_once()
        mock_get_dim.assert_called_once()
        mock_read.assert_called_once()
        mock_extract.assert_called_once()
        mock_complete.assert_called_once()
        mock_process_column.assert_called_once()
        
        # Verify the result is as expected
        self.assertEqual(result.equals(expected_result), True)


if __name__ == '__main__':
    unittest.main()