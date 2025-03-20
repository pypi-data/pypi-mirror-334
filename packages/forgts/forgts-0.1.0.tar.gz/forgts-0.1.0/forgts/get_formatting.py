import pandas as pd
import openpyxl
from openpyxl.styles import colors
from typing import Union, Dict, Any, Optional


"""
A class to handle Excel file formatting extraction and processing.

This class provides methods to load an Excel workbook, determine worksheet dimensions,
read and validate spreadsheet data, extract cell formatting, and process column formatting.
It supports handling both sheet names and indices, and ensures all cells have formatting
records by filling in missing entries.

Attributes:
    xlfilepath (str): Path to the Excel file.
    sheet (Union[str, int, None]): Sheet name or index to process. Defaults to the first sheet.
    workbook (openpyxl.Workbook): Loaded Excel workbook.
    worksheet (openpyxl.worksheet.worksheet.Worksheet): Active worksheet.
    dimensions (tuple): Dimensions of the worksheet (rows, columns).
    spreadsheet_data (pd.DataFrame): DataFrame containing spreadsheet data.
    formatting_data (pd.DataFrame): DataFrame containing formatting information.

Methods:
    load_workbook(): Load the Excel workbook and the specified worksheet.
    get_dimensions(): Determine the dimensions of the worksheet based on non-empty cells.
    read_and_validate_spreadsheet(): Read the spreadsheet data and validate its structure.
    extract_cell_formatting(): Extract formatting information for all cells in the worksheet.
    extract_single_cell_formatting(cell, row, col): Extract formatting details from a single cell.
    get_color(color): Extract color value from openpyxl Color objects.
    complete_formatting_data(): Ensure all cells have formatting records by filling in missing entries.
    create_empty_cell_record(row, col): Create a record for an empty cell with default formatting values.
    process_column_formatting(): Process formatting for each column and combine into a single DataFrame.
    process_single_column_formatting(column_name): Process formatting for a single column.
    run(): Execute the entire process of extracting and processing formatting data.
"""


class ExcelFormatter:
    def __init__(self, xlfilepath: str, sheet: Union[str, int, None] = None):
        """
        Initialize the Excel formatting handler with the file path and sheet identifier.

        Args:
            xlfilepath (str): Path to the Excel file.
            sheet (Union[str, int, None], optional): Sheet name or index to process. Defaults to the first sheet.
        """
        self.xlfilepath = xlfilepath
        self.sheet = sheet if sheet is not None else 0  # Default to the first sheet
        self.workbook = None
        self.worksheet = None
        self.dimensions = None
        self.spreadsheet_data = None
        self.formatting_data = None

    def load_workbook(self):
        """
        Load the Excel workbook and set the active worksheet.

        This method loads the Excel workbook from the specified file path and sets the
        active worksheet based on the provided sheet name or index.

        Raises:
            KeyError: If the specified sheet name does not exist in the workbook.
            IndexError: If the specified sheet index is out of range.
        """
        self.workbook = openpyxl.load_workbook(self.xlfilepath)
        if isinstance(self.sheet, int):
            self.worksheet = self.workbook.worksheets[self.sheet]
        else:
            self.worksheet = self.workbook[self.sheet]

    def get_dimensions(self):
        """
        Determine the dimensions of the worksheet based on non-empty cells.

        Iterates through all cells in the worksheet to find the last row and column
        containing data. If no data is found, defaults to the worksheet's maximum
        row and column values. Updates the `dimensions` attribute with the determined
        dimensions.

        Updates:
            self.dimensions (tuple): The dimensions of the worksheet (rows, columns).
        """
        last_row = 0
        last_col = 0

        for row_idx, row in enumerate(self.worksheet.iter_rows(), 1):
            for col_idx, cell in enumerate(row, 1):
                if cell.value is not None:  # Only consider cells with actual values
                    last_row = max(last_row, row_idx)
                    last_col = max(last_col, col_idx)

        # Fall back to worksheet dimensions if no data found
        if last_row == 0 or last_col == 0:
            last_row = self.worksheet.max_row
            last_col = self.worksheet.max_column

        self.dimensions = (last_row, last_col)

    def read_and_validate_spreadsheet(self):
        """
        Reads the spreadsheet data from the specified Excel file and validates its structure.

        Loads the data into a DataFrame, treating empty cells and empty strings as NaN.
        Raises a ValueError if any column names are detected as empty, indicated by names
        starting with '...'.

        Raises:
            ValueError: If the spreadsheet contains empty values in the header row.
        """
        self.spreadsheet_data = pd.read_excel(
            self.xlfilepath,
            sheet_name=self.sheet,
            keep_default_na=True,  # Treats empty cells as NaN
            na_values=[""],  # Explicitly treat empty strings as NaN
        )

        # Check for empty column names
        if any(str(name).startswith("...") for name in self.spreadsheet_data.columns):
            raise ValueError("Check the spreadsheet for empty values in the header row")

    def extract_cell_formatting(self):
        """
        Extract formatting information for all cells in the worksheet.

        Iterates over each cell in the worksheet, extracts its formatting details,
        and stores the information in a DataFrame. The extracted data includes
        font styles, border styles, and colors.

        Updates:
            self.formatting_data (pd.DataFrame): DataFrame containing formatting
            information for each cell.
        """
        formatting_records = []
        last_row, last_col = self.dimensions

        for row in range(1, last_row + 1):
            for col in range(1, last_col + 1):
                cell = self.worksheet.cell(row=row, column=col)
                cell_data = self.extract_single_cell_formatting(cell, row, col)
                formatting_records.append(cell_data)

        self.formatting_data = pd.DataFrame(formatting_records)

    @staticmethod
    def extract_single_cell_formatting(
        cell: openpyxl.cell.cell.Cell, row: int, col: int
    ) -> Dict[str, Any]:
        font = cell.font
        fill = cell.fill
        border = cell.border
        """
        Extract formatting details from a single cell.

        Args:
            cell (openpyxl.cell.cell.Cell): The cell from which to extract formatting.
            row (int): The row number of the cell.
            col (int): The column number of the cell.

        Returns:
            Dict[str, Any]: A dictionary containing the cell's formatting details,
            including font styles, border styles, and colors.
        """
        return {
            "row": row,
            "col": col,
            "bold": font.bold,
            "italic": font.italic,
            "underlined": font.underline if font.underline else False,
            "strikethrough": font.strike if font.strike else False,
            "border_top_style": border.top.style if border.top else None,
            "border_right_style": border.right.style if border.right else None,
            "border_bottom_style": border.bottom.style if border.bottom else None,
            "border_left_style": border.left.style if border.left else None,
            "hl_color": (
                ExcelFormatter.get_color(fill.start_color) if fill.start_color else None
            ),
            "text_clr": ExcelFormatter.get_color(font.color) if font.color else None,
            "border_top_clr": (
                ExcelFormatter.get_color(border.top.color)
                if border.top and border.top.color
                else None
            ),
            "border_right_clr": (
                ExcelFormatter.get_color(border.right.color)
                if border.right and border.right.color
                else None
            ),
            "border_bottom_clr": (
                ExcelFormatter.get_color(border.bottom.color)
                if border.bottom and border.bottom.color
                else None
            ),
            "border_left_clr": (
                ExcelFormatter.get_color(border.left.color)
                if border.left and border.left.color
                else None
            ),
        }

    @staticmethod
    def get_color(color) -> Optional[str]:
        """
        Extracts the color value from an openpyxl Color object.

        Args:
            color: An openpyxl Color object representing the cell's color.

        Returns:
            A string representing the color in RGB format, theme index, or 'auto',
            or None if the color is not defined or recognized.
        """
        if color is None:
            return None

        if color.type == "rgb":
            return color.rgb.upper()

        if color.type == "indexed":
            if color.indexed < len(colors.COLOR_INDEX):
                indexed_rgb = colors.COLOR_INDEX[color.indexed]
                return indexed_rgb.upper()

        if color.type == "theme":
            return f"{color.theme}"

        if color.type == "auto":
            return "auto"

        return None

    def complete_formatting_data(self):
        """
        Ensure all rows in the worksheet have corresponding formatting records.

        This method checks for any missing rows in the formatting data and creates
        default formatting records for them. It iterates over each unique column
        value, identifies missing rows, and appends default records for these rows
        using `create_empty_cell_record`. The complete set of formatting data is
        then consolidated into a single DataFrame.

        Modifies:
            self.formatting_data (pd.DataFrame): Updated to include all rows with
            default formatting for any previously missing entries.
        """
        max_row = self.formatting_data["row"].max()
        all_rows = set(range(1, max_row + 1))

        complete_data_frames = []

        for col_val in self.formatting_data["col"].unique():
            col_data = self.formatting_data[self.formatting_data["col"] == col_val]
            existing_rows = set(col_data["row"])
            missing_rows = all_rows - existing_rows

            # Add existing rows
            complete_data_frames.append(col_data)

            # Add missing rows if any
            if missing_rows:
                missing_data = [
                    self.create_empty_cell_record(row, col_val) for row in missing_rows
                ]
                complete_data_frames.append(pd.DataFrame(missing_data))

        # Combine all data
        self.formatting_data = pd.concat(complete_data_frames, ignore_index=True)

    @staticmethod
    def create_empty_cell_record(row: int, col: int) -> Dict[str, Optional[Any]]:
        """
        Create a record for an empty cell with default formatting values.

        Args:
            row (int): The row index of the cell.
            col (int): The column index of the cell.

        Returns:
            Dict[str, Optional[Any]]: A dictionary containing default formatting attributes
            for the cell, with all values set to None.
        """
        return {
            "row": row,
            "col": col,
            "bold": None,
            "italic": None,
            "underlined": None,
            "hl_color": None,
            "strikethrough": None,
            "text_clr": None,
            "border_top_style": None,
            "border_top_clr": None,
            "border_right_style": None,
            "border_right_clr": None,
            "border_bottom_style": None,
            "border_bottom_clr": None,
            "border_left_style": None,
            "border_left_clr": None,
        }

    def process_column_formatting(self):
        """
        Process formatting for all columns in the spreadsheet.

        Iterates over each column in the spreadsheet data, processes its formatting
        using `process_single_column_formatting`, and concatenates the results into
        a single DataFrame.

        Returns:
            pd.DataFrame: A DataFrame containing the combined formatting information
            for all columns.
        """
        column_formatting_frames = [
            self.process_single_column_formatting(col_name)
            for col_name in self.spreadsheet_data.columns
        ]

        return pd.concat(column_formatting_frames, ignore_index=True)

    def process_single_column_formatting(self, column_name: str) -> pd.DataFrame:
        """
        Process the formatting of a single column in the spreadsheet.

        This method extracts and processes the formatting details for a specified column
        from the spreadsheet data. It filters the formatting data for the given column,
        selects relevant formatting attributes, and reshapes the data into a long format
        DataFrame suitable for further analysis or reporting.

        Args:
            column_name (str): The name of the column to process.

        Returns:
            pd.DataFrame: A DataFrame containing the processed formatting details for the
            specified column, with each formatting attribute represented as a separate row.
        """
        column_index = list(self.spreadsheet_data.columns).index(column_name) + 1

        column_format = self.formatting_data[
            (self.formatting_data["row"] >= 2)
            & (self.formatting_data["col"] == column_index)
        ].copy()

        format_columns = [
            "bold",
            "italic",
            "underlined",
            "hl_color",
            "strikethrough",
            "text_clr",
            "border_top_style",
            "border_top_clr",
            "border_right_style",
            "border_right_clr",
            "border_bottom_style",
            "border_bottom_clr",
            "border_left_style",
            "border_left_clr",
        ]
        column_format = column_format[format_columns].copy()

        column_format.insert(0, "target_var", column_name)
        column_format.insert(0, "rowid", range(1, len(column_format) + 1))

        column_format = column_format.astype(str)

        return pd.melt(
            column_format,
            id_vars=["rowid", "target_var"],
            var_name="format",
            value_name="val",
        )

    def run(self):
        """
        Execute the sequence of operations to process the Excel spreadsheet.

        This method orchestrates the loading of the workbook, determining its dimensions,
        validating the spreadsheet data, extracting cell formatting, completing formatting
        data, and finally processing column formatting.

        Returns:
            pd.DataFrame: A DataFrame containing the combined formatting information
            for all columns.
        """
        self.load_workbook()
        self.get_dimensions()
        self.read_and_validate_spreadsheet()
        self.extract_cell_formatting()
        self.complete_formatting_data()
        return self.process_column_formatting()
