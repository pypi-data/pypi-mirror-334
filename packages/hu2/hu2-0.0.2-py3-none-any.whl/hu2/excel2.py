from typing import Dict
from openpyxl.workbook import Workbook

class ExcelTable:
    def __init__(self,file_name):
        self.head_list = None
        self.body_grid = None
        self.file_name = file_name
        self.wb = Workbook()
        self.sheet = self.wb.active
        

    def set_head(self,head_list:list|tuple):
        self.head_list = head_list
    
    def set_body(self, body_grid):
        self.body_grid = body_grid
    
    def set_column_width(self, col_width_dict:Dict[str,int]):
        for col_str,col_width in col_width_dict.items():
            self.sheet.column_dimensions[col_str].width = col_width
        
    
    def save(self):
        if self.head_list is not None:
            for index, head in enumerate(self.head_list):
                self.sheet.cell(row=1,column=index+1,value=str(head))

        if self.body_grid is not None:
            for rindex in range(len(self.body_grid)):
                for colindex in range(len(self.body_grid[rindex])):
                    self.sheet.cell(row=rindex+2,column=colindex+1,value=self.body_grid[rindex][colindex])

        self.wb.save(self.file_name)



if __name__ == '__main__':
    print('ok')
    et = ExcelTable('excel.xlsx')
    et.set_column_width({'A':10,'B':20,'C':30})
    et.set_head(['Aa','Bb','Cc'])
    et.set_body([['a','b','c'],(1,2,3)])

    et.save()

    #A1  = 