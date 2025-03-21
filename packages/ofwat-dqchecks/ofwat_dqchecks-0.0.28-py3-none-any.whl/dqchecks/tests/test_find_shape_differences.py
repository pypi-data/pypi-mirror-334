"""
Test function related to find_shape_differences
function in panacea.py file
"""
import pytest
import pandas as pd
from openpyxl import Workbook
from dqchecks.panacea import (
    create_dataframe_structure_discrepancies,
    find_shape_differences,
    StructureDiscrepancyContext)

def test_create_dataframe_valid_input():
    """Valid input data and context"""
    input_data = {
        'errors': {
            'row_count_discrepancy': ["Row count mismatch", "Rows missing"],
            'column_count_discrepancy': ["Column count mismatch"]
        }
    }
    context = StructureDiscrepancyContext(
        Rule_Cd="123",
        Sheet_Cd="Sheet1",
        Error_Category="Structure Discrepancy",
        Error_Severity_Cd="high",
    )

    df = create_dataframe_structure_discrepancies(input_data, context)

    assert isinstance(df, pd.DataFrame)
    assert len(df) == 2  # 2 since lists are concatenated. Could make it 3
    assert set(df.columns) == {
        'Event_Id',
        'Sheet_Cd',
        'Rule_Cd',
        'Error_Category',
        'Error_Severity_Cd',
        'Error_Desc'}
    assert df['Sheet_Cd'][0] == "Sheet1"
    assert df['Error_Desc'][0] == "Row count mismatch -- Rows missing"

def test_create_dataframe_invalid_input_data():
    """Invalid input data (not a dictionary)"""
    input_data = []  # This is an invalid input (not a dictionary)
    context = StructureDiscrepancyContext(
        Rule_Cd="123",
        Sheet_Cd="Sheet1",
        Error_Category="Structure Discrepancy",
        Error_Severity_Cd="high",
    )

    with pytest.raises(TypeError):
        create_dataframe_structure_discrepancies(input_data, context)

def test_create_dataframe_invalid_context():
    """Invalid context (not an instance of StructureDiscrepancyContext)"""
    input_data = {
        'errors': {
            'row_count_discrepancy': ["Row count mismatch"]
        }
    }
    context = {}  # This is an invalid context (not an instance of StructureDiscrepancyContext)

    with pytest.raises(TypeError):
        create_dataframe_structure_discrepancies(input_data, context)

def test_create_dataframe_missing_errors_field():
    """Missing 'errors' field in input data"""
    input_data = {}  # Missing 'errors' field
    context = StructureDiscrepancyContext(
        Rule_Cd="123",
        Sheet_Cd="Sheet1",
        Error_Category="Structure Discrepancy",
        Error_Severity_Cd="high",
    )

    with pytest.raises(ValueError):
        create_dataframe_structure_discrepancies(input_data, context)

def test_create_dataframe_invalid_discrepancy_format():
    """Discrepancy that is not a list or tuple"""
    input_data = {
        'errors': {
            'row_count_discrepancy': "Row count mismatch"  # Not a list or tuple
        }
    }
    context = StructureDiscrepancyContext(
        Rule_Cd="123",
        Sheet_Cd="Sheet1",
        Error_Category="Structure Discrepancy",
        Error_Severity_Cd="high",
    )

    with pytest.raises(ValueError):
        create_dataframe_structure_discrepancies(input_data, context)

def test_create_dataframe_invalid_discrepancy_element():
    """Discrepancy list contains non-string elements"""
    input_data = {
        'errors': {
            'row_count_discrepancy': [123, "Row count mismatch"]  # Non-string element
        }
    }
    context = StructureDiscrepancyContext(
        Rule_Cd="123",
        Sheet_Cd="Sheet1",
        Error_Category="Structure Discrepancy",
        Error_Severity_Cd="high",
    )

    with pytest.raises(ValueError):
        create_dataframe_structure_discrepancies(input_data, context)

def test_create_dataframe_missing_context_attributes():
    """ Context missing one or more attributes"""
    input_data = {
        'errors': {
            'row_count_discrepancy': ["Row count mismatch"]
        }
    }
    # Missing 'Rule_Cd' in context
    context = StructureDiscrepancyContext(
        Rule_Cd=None,
        Sheet_Cd="Sheet1",
        Error_Category="Structure Discrepancy",
        Error_Severity_Cd="high",
    )

    with pytest.raises(ValueError):
        create_dataframe_structure_discrepancies(input_data, context)

def test_create_dataframe_empty_errors_field():
    """Edge case with empty 'errors' field"""
    input_data = {
        'errors': {}
    }
    context = StructureDiscrepancyContext(
        Rule_Cd="123",
        Sheet_Cd="Sheet1",
        Error_Category="Structure Discrepancy",
        Error_Severity_Cd="high",
    )

    df = create_dataframe_structure_discrepancies(input_data, context)

    assert isinstance(df, pd.DataFrame)
    assert df.empty  # The DataFrame should be empty

def test_create_dataframe_multiple_error_types():
    """Multiple error types and discrepancies"""
    input_data = {
        'errors': {
            'row_count_discrepancy': ["Row count mismatch", "Rows missing"],
            'column_count_discrepancy': ["Column count mismatch", "Columns missing"]
        }
    }
    context = StructureDiscrepancyContext(
        Rule_Cd="123",
        Sheet_Cd="Sheet1",
        Error_Category="Structure Discrepancy",
        Error_Severity_Cd="high",
    )

    df = create_dataframe_structure_discrepancies(input_data, context)

    assert len(df) == 2  # Two as two error categories
    assert df['Error_Desc'][0] == "Row count mismatch -- Rows missing"
    assert df['Error_Desc'][1] == "Column count mismatch -- Columns missing"

def test_find_shape_differences_with_discrepancies():
    """Test when both workbooks have common sheets with discrepancies"""
    wb_template = Workbook()
    wb_company = Workbook()

    # Create a sheet in each workbook
    sheet_template = wb_template.create_sheet("Sheet1")
    sheet_company = wb_company.create_sheet("Sheet1")

    del wb_template["Sheet"]
    del wb_company["Sheet"]

    # Add data to simulate discrepancies in structure
    sheet_template['A1'] = "Header1"
    sheet_template['A2'] = "Data1"
    sheet_company['A1'] = "Header1"
    sheet_company['A2'] = "Data2"
    sheet_company['B1'] = "ExtraColumn"  # This extra column will cause a discrepancy

    # Run the function
    result_df = find_shape_differences(wb_template, wb_company)

    # Check if the returned DataFrame has discrepancies
    assert not result_df.empty
    assert len(result_df) > 0  # Ensure there is at least one discrepancy
    assert set(result_df['Sheet_Cd'].to_list()) == set(["Sheet1", "Sheet1", "Sheet1"])
    assert set(result_df['Rule_Cd'].to_list()) == set(["?", "?", "?"])
    assert set(result_df['Error_Category'].to_list()) == set(
        ['Structure Discrepancy', 'Structure Discrepancy', 'Structure Discrepancy'])
    assert set(result_df['Error_Severity_Cd'].to_list()) == set(
        ["hard", "hard", "hard"])
    assert set(result_df['Error_Desc'].to_list()) == set(
        ['Sheet1', "'Sheet1' has 2 rows and 1 columns, 'Sheet1' has 2 rows and 2 columns.", ''])

def test_find_shape_differences_no_common_sheets():
    """Test when both workbooks have no common sheets"""
    wb_template = Workbook()
    wb_company = Workbook()

    del wb_template["Sheet"]
    del wb_company["Sheet"]

    # Create different sheets in each workbook
    wb_template.create_sheet("Sheet1")
    wb_company.create_sheet("Sheet2")

    # Run the function
    result_df = find_shape_differences(wb_template, wb_company)

    # Check if the returned DataFrame is empty (no common sheets)
    assert result_df.empty

def test_find_shape_differences_invalid_workbook_type():
    """Test when invalid workbook types are passed"""
    with pytest.raises(TypeError):
        find_shape_differences("invalid_template", "invalid_company")
