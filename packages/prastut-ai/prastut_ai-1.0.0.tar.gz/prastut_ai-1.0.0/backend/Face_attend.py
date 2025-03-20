"""
Face Recognition Attendance System
--------------------------------
This module implements a real-time face recognition system for attendance tracking.
It uses MTCNN for face detection, FaceNet for embedding generation, and integrates
with Pinecone for vector similarity search.

Author: Prastut AI
Version: 1.0.0
"""

import cv2 as cv
import time
import pickle
import os
import numpy as np
import openpyxl as op
import datetime as dt
from concurrent.futures import ThreadPoolExecutor
import logging
from deepface import DeepFace
from facenet_pytorch import MTCNN, InceptionResnetV1
import torch
import pinecone
from dotenv import load_dotenv
import firebase_admin
from firebase_admin import credentials, firestore
import json
from flask import Flask, jsonify, request, render_template, redirect
from werkzeug.serving import WSGIServer

# Load environment variables
load_dotenv()

# Configure logging with timestamp and log level
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - [%(filename)s:%(lineno)d] - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

def init_firebase():
    """Initialize Firebase connection with error handling."""
    try:
        cred_path = os.getenv('FIREBASE_CREDENTIALS')
        if not cred_path:
            logger.error("Firebase credentials path not found in environment variables")
            raise ValueError("FIREBASE_CREDENTIALS environment variable not set")
            
        cred = credentials.Certificate(cred_path)
        firebase_admin.initialize_app(cred)
        return firestore.client()
    except Exception as e:
        logger.error(f"Failed to initialize Firebase: {str(e)}")
        raise

def init_pinecone():
    """Initialize Pinecone connection with error handling."""
    try:
        api_key = os.getenv('PINECONE_API_KEY')
        env = os.getenv('PINECONE_ENV')
        
        if not api_key or not env:
            logger.error("Pinecone credentials not found in environment variables")
            raise ValueError("Pinecone environment variables not set")
            
        pinecone.init(api_key=api_key, environment=env)
        
        index_name = "face-embeddings"
        if index_name not in pinecone.list_indexes():
            logger.info(f"Creating new Pinecone index: {index_name}")
            pinecone.create_index(index_name, dimension=512, metric="cosine")
            
        return pinecone.Index(index_name)
    except Exception as e:
        logger.error(f"Failed to initialize Pinecone: {str(e)}")
        raise

# Initialize Flask
app = Flask(__name__)

class AttendanceManager:
    """
    Manages attendance records using Excel spreadsheets.
    Handles record creation, updates, and report generation.
    """
    
    def __init__(self):
        """Initialize AttendanceManager with the attendance file."""
        self.filename = "attendance_record.xlsx"
        self.initialize_workbook()
        
    def initialize_workbook(self):
        """Create or load the attendance workbook."""
        try:
            if not os.path.exists(self.filename):
                logger.info("Creating new attendance workbook")
                wb = op.Workbook()
                ws = wb.active
                ws.title = "Attendance"
                
                headers = ["Date", "Name", "Roll Number", "Time", "Confidence"]
                for col, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col, value=header)
                wb.save(self.filename)
        except Exception as e:
            logger.error(f"Failed to initialize workbook: {str(e)}")
            raise

    def add_attendance(self, name, roll_number, confidence):
        """
        Add an attendance record with error handling and duplicate checking.
        
        Args:
            name (str): Student name
            roll_number (str): Student roll number
            confidence (float): Face recognition confidence score
            
        Returns:
            tuple: (success: bool, message: str)
        """
        try:
            current_time = dt.datetime.now()
            current_date = current_time.strftime("%Y-%m-%d")
            time_str = current_time.strftime("%H:%M:%S")

            # Load workbook
            try:
                wb = op.load_workbook(self.filename)
            except Exception as e:
                logger.error(f"Failed to load attendance workbook: {str(e)}")
                return False, "Failed to load attendance records"

            ws = wb.active

            # Check for duplicate entry
            for row in range(2, ws.max_row + 1):
                if (ws.cell(row=row, column=1).value == current_date and 
                    ws.cell(row=row, column=2).value == name):
                    logger.info(f"Duplicate attendance entry for {name} on {current_date}")
                    return False, f"Attendance already marked for {name} today"

            # Add new record
            try:
                next_row = ws.max_row + 1
                ws.cell(row=next_row, column=1, value=current_date)
                ws.cell(row=next_row, column=2, value=name)
                ws.cell(row=next_row, column=3, value=roll_number)
                ws.cell(row=next_row, column=4, value=time_str)
                ws.cell(row=next_row, column=5, value=confidence)

                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = op.utils.get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    ws.column_dimensions[column_letter].width = adjusted_width

                wb.save(self.filename)
                logger.info(f"Attendance marked successfully for {name}")
                return True, "Attendance marked successfully"
                
            except Exception as e:
                logger.error(f"Failed to write attendance record: {str(e)}")
                return False, "Failed to save attendance record"

        except Exception as e:
            logger.error(f"Error in add_attendance: {str(e)}")
            return False, f"Error marking attendance: {str(e)}"

    def get_daily_report(self, date=None):
        """
        Generate attendance report for a specific date.
        
        Args:
            date (str, optional): Date in YYYY-MM-DD format. Defaults to today.
            
        Returns:
            list: List of attendance records for the specified date
        """
        try:
            if date is None:
                date = dt.datetime.now().strftime("%Y-%m-%d")

            wb = op.load_workbook(self.filename)
            ws = wb.active
            
            report = []
            for row in range(2, ws.max_row + 1):
                if ws.cell(row=row, column=1).value == date:
                    report.append({
                        "name": ws.cell(row=row, column=2).value,
                        "roll_number": ws.cell(row=row, column=3).value,
                        "time": ws.cell(row=row, column=4).value,
                        "confidence": ws.cell(row=row, column=5).value
                    })
            return report

        except Exception as e:
            logger.error(f"Error generating report: {str(e)}")
            return []

class FaceRecognitionSystem:
    """
    Core face recognition system that handles real-time face detection,
    embedding generation, and attendance tracking.
    
    Attributes:
        mtcnn: MTCNN model for face detection
        facenet: FaceNet model for embedding generation
        attendance_manager: Instance of AttendanceManager
    """
    
    def __init__(self):
        """Initialize face recognition models and dependencies."""
        try:
            # Initialize MTCNN for face detection
            logger.info("Initializing MTCNN model...")
            self.mtcnn = MTCNN(
                image_size=160,
                margin=0,
                min_face_size=20,
                thresholds=[0.6, 0.7, 0.7],
                factor=0.709,
                post_process=True,
                device='cuda' if torch.cuda.is_available() else 'cpu'
            )
            
            # Initialize FaceNet model
            logger.info("Initializing FaceNet model...")
            self.facenet = InceptionResnetV1(pretrained='vggface2').eval()
            if torch.cuda.is_available():
                logger.info("CUDA available - moving models to GPU")
                self.facenet = self.facenet.cuda()
            
            self.attendance_manager = AttendanceManager()
            logger.info("Face recognition system initialized successfully")
            
        except Exception as e:
            logger.error(f"Failed to initialize face recognition system: {str(e)}")
            raise
            
    def get_face_embedding(self, face_img):
        """
        Generate face embedding using FaceNet.
        
        Args:
            face_img: numpy array of face image
            
        Returns:
            numpy array: 512-dimensional face embedding vector
        """
        try:
            # Preprocess face image
            face_tensor = self.mtcnn(face_img)
            if face_tensor is None:
                logger.warning("No face detected in image")
                return None
                
            # Move to GPU if available
            if torch.cuda.is_available():
                face_tensor = face_tensor.cuda()
                
            # Generate embedding
            with torch.no_grad():
                embedding = self.facenet(face_tensor.unsqueeze(0))
            return embedding.detach().cpu().numpy()[0]
            
        except Exception as e:
            logger.error(f"Error generating face embedding: {str(e)}")
            return None
            
    def process_frame(self, frame):
        """
        Process a single frame and detect faces.
        
        Args:
            frame: numpy array of input frame
            
        Returns:
            list: List of dictionaries containing detection results
        """
        try:
            # Detect faces using MTCNN
            boxes, _ = self.mtcnn.detect(frame)
            
            if boxes is None:
                return []
                
            results = []
            for box in boxes[:20]:  # Limit to 20 people per frame
                try:
                    x1, y1, x2, y2 = [int(b) for b in box]
                    
                    # Extract face region
                    face_img = frame[y1:y2, x1:x2]
                    
                    # Get face embedding
                    embedding = self.get_face_embedding(face_img)
                    if embedding is None:
                        logger.warning("Failed to generate embedding for detected face")
                        continue
                        
                    # Query Pinecone for similar faces
                    query_result = face_index.query(
                        vector=embedding.tolist(),
                        top_k=1,
                        include_metadata=True
                    )
                    
                    # Process match results
                    if query_result.matches and query_result.matches[0].score > 0.7:
                        match = query_result.matches[0]
                        name = match.metadata.get('name', 'Unknown')
                        confidence = float(match.score) * 100
                        
                        # Verify with DeepFace for additional accuracy
                        try:
                            verification = DeepFace.verify(
                                face_img,
                                match.metadata.get('image_path'),
                                model_name="Facenet",
                                enforce_detection=False
                            )
                            if verification['verified']:
                                results.append({
                                    'box': (x1, y1, x2-x1, y2-y1),
                                    'name': name,
                                    'confidence': confidence,
                                    'verified': True
                                })
                                
                                # Mark attendance
                                success, message = self.attendance_manager.add_attendance(
                                    name,
                                    match.metadata.get('roll_number', 'NA'),
                                    confidence
                                )
                                if not success:
                                    logger.warning(f"Failed to mark attendance: {message}")
                                    
                        except Exception as e:
                            logger.warning(f"DeepFace verification failed: {str(e)}")
                            
                except Exception as e:
                    logger.error(f"Error processing individual face: {str(e)}")
                    continue
                        
            return results
            
        except Exception as e:
            logger.error(f"Error processing frame: {str(e)}")
            return []

    def run_detection(self):
        """Run the face detection system with real-time video capture."""
        try:
            cap = cv.VideoCapture(0)
            cap.set(cv.CAP_PROP_FRAME_WIDTH, 1920)
            cap.set(cv.CAP_PROP_FRAME_HEIGHT, 1080)
            
            if not cap.isOpened():
                logger.error("Could not open video capture device")
                return
                
            logger.info("Starting enhanced face recognition system...")
            print("Press 'x' to exit")
            
            while True:
                try:
                    ret, frame = cap.read()
                    if not ret:
                        logger.error("Failed to read frame from camera")
                        break
                        
                    # Process frame
                    results = self.process_frame(frame)
                    
                    # Draw results
                    for result in results:
                        x, y, w, h = result['box']
                        confidence = result['confidence']
                        
                        # Draw rectangle and name
                        color = (0, 255, 0) if result['verified'] else (0, 0, 255)
                        cv.rectangle(frame, (x, y), (x+w, y+h), color, 2)
                        cv.putText(frame, f"{result['name']} {confidence:.1f}%",
                                  (x, y-10), cv.FONT_HERSHEY_SIMPLEX, 0.9, color, 2)
                                  
                        # Draw zoomed face
                        face = frame[y:y+h, x:x+w]
                        if face.size > 0:
                            face = cv.resize(face, (100, 100))
                            frame[10:110, 10:110] = face
                    
                    # Add info overlay
                    cv.putText(frame, f"Detected: {len(results)}/20",
                              (frame.shape[1]-200, 30), cv.FONT_HERSHEY_SIMPLEX,
                              1, (0, 255, 0), 2)
                              
                    # Display frame
                    cv.imshow("Enhanced Face Recognition System", frame)
                    
                    if cv.waitKey(1) & 0xFF == ord('x'):
                        break
                        
                except Exception as e:
                    logger.error(f"Error in main detection loop: {str(e)}")
                    continue
                    
        except Exception as e:
            logger.error(f"Error in run_detection: {str(e)}")
            
        finally:
            cap.release()
            cv.destroyAllWindows()

# Initialize global instances
try:
    logger.info("Initializing system components...")
    db = init_firebase()
    face_index = init_pinecone()
except Exception as e:
    logger.error(f"Failed to initialize system components: {str(e)}")
    raise

# Flask routes for API
@app.route('/api/register', methods=['POST'])
def register_face():
    """
    Register a new face in the system.
    
    Request body:
        name (str): Person's name
        roll_number (str): Roll number/ID
        image_path (str): Path to face image
        
    Returns:
        JSON response with success/error message
    """
    try:
        data = request.json
        required_fields = ['name', 'roll_number', 'image_path']
        
        # Validate request data
        for field in required_fields:
            if field not in data:
                logger.error(f"Missing required field: {field}")
                return jsonify({'error': f"Missing required field: {field}"}), 400
                
        name = data['name']
        roll_number = data['roll_number']
        image_path = data['image_path']
        
        # Validate image path
        if not os.path.exists(image_path):
            logger.error(f"Image not found: {image_path}")
            return jsonify({'error': 'Image file not found'}), 404
        
        # Initialize face recognition system
        face_system = FaceRecognitionSystem()
        
        # Generate embedding
        image = cv.imread(image_path)
        if image is None:
            logger.error(f"Failed to read image: {image_path}")
            return jsonify({'error': 'Failed to read image file'}), 400
            
        embedding = face_system.get_face_embedding(image)
        
        if embedding is None:
            return jsonify({'error': 'No face detected in image'}), 400
            
        # Store in Pinecone
        try:
            face_index.upsert(
                vectors=[{
                    'id': f"{name}_{roll_number}",
                    'values': embedding.tolist(),
                    'metadata': {
                        'name': name,
                        'roll_number': roll_number,
                        'image_path': image_path
                    }
                }]
            )
        except Exception as e:
            logger.error(f"Failed to store face embedding in Pinecone: {str(e)}")
            return jsonify({'error': 'Failed to store face data'}), 500
        
        # Store in Firebase
        try:
            db.collection('students').document(f"{name}_{roll_number}").set({
                'name': name,
                'roll_number': roll_number,
                'image_path': image_path,
                'registered_at': firestore.SERVER_TIMESTAMP
            })
        except Exception as e:
            logger.error(f"Failed to store user data in Firebase: {str(e)}")
            return jsonify({'error': 'Failed to store user data'}), 500
        
        logger.info(f"Successfully registered face for {name}")
        return jsonify({'message': 'Face registered successfully'}), 200
        
    except Exception as e:
        logger.error(f"Error registering face: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/attendance', methods=['GET'])
def get_attendance():
    """
    Get attendance records for a specific date.
    
    Query parameters:
        date (str, optional): Date in YYYY-MM-DD format
        
    Returns:
        JSON response with attendance records
    """
    try:
        date = request.args.get('date')
        
        # Validate date format if provided
        if date:
            try:
                dt.datetime.strptime(date, '%Y-%m-%d')
            except ValueError:
                logger.error(f"Invalid date format: {date}")
                return jsonify({'error': 'Invalid date format. Use YYYY-MM-DD'}), 400
        
        attendance_manager = AttendanceManager()
        report = attendance_manager.get_daily_report(date)
        
        logger.info(f"Retrieved {len(report)} attendance records for {date or 'today'}")
        return jsonify(report), 200
        
    except Exception as e:
        logger.error(f"Error getting attendance: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/')
def trial_page():
    """Render the trial offer page."""
    try:
        return render_template('trial_offer.html')
    except Exception as e:
        logger.error(f"Error rendering trial page: {str(e)}")
        return jsonify({'error': 'Failed to load trial page'}), 500

@app.route('/submit_trial', methods=['POST'])
def submit_trial():
    """
    Handle trial form submission.
    
    Form data:
        school_name (str): Name of the school
        contact_person (str): Contact person's name
        email (str): Contact email
        phone (str): Contact phone number
        student_count (str): Number of students
    """
    try:
        data = request.form
        required_fields = ['school_name', 'contact_person', 'email', 'phone']
        
        # Validate form data
        for field in required_fields:
            if not data.get(field):
                logger.error(f"Missing required field in trial form: {field}")
                return jsonify({'error': f"Missing required field: {field}"}), 400
        
        # Store trial request in Firebase
        try:
            doc_ref = db.collection('trial_requests').add({
                'school_name': data.get('school_name'),
                'contact_person': data.get('contact_person'),
                'email': data.get('email'),
                'phone': data.get('phone'),
                'student_count': data.get('student_count'),
                'requested_at': firestore.SERVER_TIMESTAMP
            })
            
            logger.info(f"Trial request submitted for {data.get('school_name')}")
            return jsonify({'message': 'Trial request submitted successfully'}), 200
            
        except Exception as e:
            logger.error(f"Failed to store trial request: {str(e)}")
            return jsonify({'error': 'Failed to submit trial request'}), 500
            
    except Exception as e:
        logger.error(f"Error submitting trial request: {str(e)}")
        return jsonify({'error': str(e)}), 500

def main():
    """Main function to run the system."""
    try:
        # Initialize face recognition system
        face_system = FaceRecognitionSystem()
        
        # Run in development mode
        if os.getenv('FLASK_ENV') == 'development':
            logger.info("Starting server in development mode")
            app.run(debug=True)
        else:
            # Run in production mode
            logger.info("Starting server in production mode")
            http_server = WSGIServer(('', 5000), app)
            http_server.serve_forever()
            
    except Exception as e:
        logger.error(f"System error: {str(e)}")
        raise
        
if __name__ == "__main__":
    main()
