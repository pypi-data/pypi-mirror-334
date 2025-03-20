import os
import uuid
from collections import (
    defaultdict,
)
from typing import (
    DefaultDict,
    List,
)

from openpyxl import (
    load_workbook,
)
from simple_report.converter.abstract import (
    FileConverter,
)
from simple_report.report import (
    SpreadsheetReport,
)
from simple_report.xlsx.document import (
    DocumentXLSX,
)

from kinder.core.unit.models import (
    Unit,
)

from gisdo import (
    settings as gs_settings,
)
from gisdo.collector import (
    ReportCacheMixin,
)

from .collector import (
    ReportDataCollector,
)
from .constants import (
    INDEXES,
    DataTypes,
)
from .helpers import (
    ReportRecord,
)


class IndexReportAdapter:
    """Адаптер отчёта "Выгрузка детей по показателю/тегу"."""

    def __init__(self, data: list, indexes: list):
        self.data = data
        self.indexes = indexes

    @staticmethod
    def get_unit_name(unit_id: int) -> str:
        """Получение названия организации по id."""
        return Unit.objects.values_list('name', flat=True).get(id=unit_id)

    @staticmethod
    def get_child_record(unit_name, age_cat, child_id, date_of_birth, fullname, data_type, data_id):
        """Получение namedtuple строки отчёта для данных ребёнка."""
        data_type_name = DataTypes.values.get(data_type, '')
        return ReportRecord(unit_name, age_cat, data_type_name, data_id, fullname)

    @staticmethod
    def get_group_record(group_data, group_id):
        """Получение namedtuple строки отчёта для данных группы."""
        return ReportRecord(
            group_data['doo_name'],
            '-',
            f'Группа ["{group_data["group_name"]}"]',
            group_id,
            group_data['value'],
        )

    @staticmethod
    def get_xml_tag_record(unit_name, value):
        """Получение namedtuple строки отчёта для данных группы."""
        return ReportRecord(unit_name, '', '', '', value)

    def get_data(self) -> DefaultDict[str, List[ReportRecord]]:
        """Получение преобразованных данных."""
        final_data = defaultdict(list)
        groups_data = list()

        for mo_data in self.data:
            for data_type in ReportCacheMixin.DATA_TYPE:
                for unit_id, index_data in mo_data[data_type].items():
                    unit_name = self.get_unit_name(unit_id)
                    if isinstance(index_data, dict):
                        for index, data in index_data.items():
                            child_records = [
                                self.get_child_record(unit_name, age_cat, *row)
                                for age_cat, rows in data.items()
                                for row in rows
                            ]
                            final_data[index].extend(child_records)
                    else:
                        groups_data.extend(index_data)

            # Отдельно обрабатываются теги из выгружаемой xml
            mo_xml_tags_data = mo_data[ReportDataCollector.XML_TAGS]
            for tag_name, tag_data in mo_xml_tags_data.items():
                for mo_id, unit_tag_value in tag_data.items():
                    for unit_id, value in unit_tag_value.items():
                        unit_name = self.get_unit_name(unit_id)
                        final_data[tag_name].append(self.get_xml_tag_record(unit_name, value))

            # Отдельно обрабатываются теги из выгружаемой xml по группам
            mo_xml_tags_data_by_group = mo_data[ReportDataCollector.XML_TAGS_BY_GROUP]
            for tag_name, tag_data in mo_xml_tags_data_by_group.items():
                for mo_id, unit_tag_value in tag_data.items():
                    for unit_id, data in unit_tag_value.items():
                        unit_name = self.get_unit_name(unit_id)
                        for age_cat, rows in data.items():
                            if rows:
                                final_data[tag_name].extend(
                                    [self.get_child_record(unit_name, age_cat, *row) for row in rows if row]
                                )
            # Отдельно обрабатываются QUEUE_8
            queue_8_data = mo_data[ReportDataCollector.QUEUE_8]
            for unit, tags_data in queue_8_data.items():
                for tag_name, tag_data in tags_data.items():
                    for age_cat, average_time in tag_data.items():
                        if average_time:
                            final_data[tag_name].append(ReportRecord(unit.name, age_cat, '', '', str(average_time)))

            # Отдельно обрабатываются теги cо значениями
            # из выгружаемой xml по группам
            mo_xml_tags_data_by_group_only_values = mo_data[ReportDataCollector.XML_TAGS_BY_GROUP_ONLY_VALUES]
            for tag_name, tag_data in mo_xml_tags_data_by_group_only_values.items():
                for mo_id, data in tag_data.items():
                    for unit_id, unit_data in data.items():
                        for group_id, group_data in unit_data.items():
                            final_data[tag_name].append(self.get_group_record(group_data, group_id))

            report_applications_data = mo_data[ReportDataCollector.REPORT_APPLICATIONS]
            for unit_id, tag_data in report_applications_data.items():
                for tag_name, unit_tag_value in tag_data.items():
                    for age_cat, value in unit_tag_value.items():
                        if value:
                            unit_name = self.get_unit_name(unit_id)
                            final_data[tag_name].append(ReportRecord(unit_name, age_cat, '', '', value))

        return final_data


class IndexReport:
    """Построитель отчета "Выгрузка детей по показателю/тегу"."""

    # Высота строк отчет по умолчанию
    DEFAULT_ROW_HEIGHT = 15
    # Номер строки, с которой начинаются данные в отчете
    DATA_ROW_NUM = 4
    # Номер колонки со значением "ФИО ребенка/Значение тега"
    TAG_COLUMN_NUM = 4

    def __init__(self, data, indexes):
        self.data = data
        self.indexes = indexes
        self.adapter = IndexReportAdapter(data, indexes)

        self.report_name = str(uuid.uuid4())[:16]

        self.template_name = os.path.join(os.path.dirname(__file__), 'index_report.xlsx')

        self.result_name = os.path.join(gs_settings.DOWNLOADS_DIR, self.report_name + '.xlsx')

        self.report = SpreadsheetReport(self.template_name, wrapper=DocumentXLSX, type=FileConverter.XLSX)

    def set_rows_height(self):
        """Установка высоты строк"""
        wb = load_workbook(filename=self.result_name)
        work_sheet = wb.worksheets[0]

        for row in work_sheet.iter_rows(min_row=self.DATA_ROW_NUM):
            row_value = row[self.TAG_COLUMN_NUM].value
            if not row_value:
                continue
            # Выставляем в высоту строк в зависимости от кол-ва строк в колонке
            # "ФИО ребенка/Значение тега"
            rows_in_cell = str(row_value).count('\n') or 1
            work_sheet.row_dimensions[row[0].row].height = self.DEFAULT_ROW_HEIGHT * rows_in_cell

        wb.save(filename=self.result_name)

    def build(self):
        """Построение отчета"""

        header_section = self.report.get_section('header')
        header_section.flush({})

        tag_section = self.report.get_section('tag')
        data_section = self.report.get_section('data')

        # Данные организации
        data = self.adapter.get_data()

        for index, index_name in INDEXES:
            if index in self.indexes:
                tag_section.flush({'tag_index': index_name})

            report_records = data.get(index)
            if not report_records:
                continue

            for report_record in report_records:
                row_data = report_record._asdict()
                data_section.flush(row_data)

        self.report.build(self.result_name)
        self.set_rows_height()

        return self.report_name
