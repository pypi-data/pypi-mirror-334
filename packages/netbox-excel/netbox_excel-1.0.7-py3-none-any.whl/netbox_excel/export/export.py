import openpyxl
from netbox_excel.models import ExportExcel
from .devices import get_device, get_device_type_parent, get_devices_child , get_device_by_id
from .rack import get_rack_have_device
import pandas as pd
# from io import StringIO
# import xlsxwriter
from django.http import HttpResponse , HttpResponseRedirect

# Vẽ tủ rack của tất cả device
def export_all_view_rack():
    # Create a new workbook and select the active worksheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Data Export"

    # create the headers
    headers = ['Rack', 'U/Rack','Tên Thiết bị', 'Chủng loại', 'Quản lý', 'Số HĐ', 'Model', 'SN', 'Thời gian lắp đặt', 'Ghi Chú']
    sheet.append(headers)

    try:
        # get device 
        devices_list = get_device()
        racks_list = [] # danh sách rack có thiết bị
        item_sheet_list = [] # danh sách item device đã xử lý dữ liệu

        # start loop for chuẩn bị dữ liệu export
        for device in devices_list:
            #Kiểm tra nếu thiết bị không nằm trong tủ rack hoặc không có position bỏ qua 
            if device.position is None:
                continue
            # get data custom feild
            device_owner = ""
            year_of_investment = ""
            contract_number = ""
            custom_fields = device.get_custom_fields_by_group()
            for key, value in custom_fields[''].items():
                if str(key) == 'Device owner' and value != None:
                    device_owner = value
                elif str(key) == 'Year of investment' and value != None:
                    year_of_investment = value
                elif str(key) == 'Contract number' and value != None:
                    contract_number = value

            # create new item export
            item_export = ExportExcel(
                rack = str(device.rack),
                device_role = str(device.role),
                device_type = str(device.device_type),
                device_name = str(device.name),
                position = int(device.position), # U/Rack
                serial_number = str(device.serial),
                device_description = str(device.description),
                owner_device = str(device_owner),
                year_of_investment = str(year_of_investment),
                contract_number = str(contract_number), 
                u_number = int(device.device_type.u_height),
            )

            # append data to item_sheet_list export
            racks_list.append(str(device.rack))
            item_sheet_list.append(item_export)
        # end loop for chuẩn bị dữ liệu export

        # lọc trùng rack 
        racks_list = sorted(list(set(racks_list)))
        u_height_sheet = 2 # đếm dòng sheet dòng 1 là header nên bắt đầu từ 2 
        # start add item into sheet 
        for rack in racks_list: 
            u_height_rack = 42 # mỗi 1 rack có 42 lần add item ( 1 item = 1 U)
            for i in range(42):
                # kiểm tra trong list item export
                device = find_device_item(item_sheet_list, rack, u_height_rack)
                if device:
                    # add item into sheet
                    item = [
                        rack,
                        u_height_rack,
                        device.device_name,
                        device.device_role,
                        device.owner_device,
                        device.contract_number,
                        device.device_type,
                        device.serial_number,
                        device.year_of_investment,
                        device.device_description,
                    ]
                    sheet.append(item)
                    
                    # check height > 1 => merg cell 
                    if device.u_number > 1:
                        # Từ cột thứ 3 đến cuối đều cần merg ô bằng chiều cao của thiết bị
                        height_device_in_sheet = u_height_sheet - device.u_number + 1
                        # copy data từ row position sang ô đầu tiên
                        for col in range(3, 11):
                            sheet.cell(row=height_device_in_sheet, column=col).value = sheet.cell(row=u_height_sheet, column=col).value
                        
                        sheet.merge_cells(start_row=height_device_in_sheet, start_column=3, end_row=u_height_sheet, end_column=3)
                        sheet.merge_cells(start_row=height_device_in_sheet, start_column=4, end_row=u_height_sheet, end_column=4)
                        sheet.merge_cells(start_row=height_device_in_sheet, start_column=5, end_row=u_height_sheet, end_column=5)
                        sheet.merge_cells(start_row=height_device_in_sheet, start_column=6, end_row=u_height_sheet, end_column=6)
                        sheet.merge_cells(start_row=height_device_in_sheet, start_column=7, end_row=u_height_sheet, end_column=7)
                        sheet.merge_cells(start_row=height_device_in_sheet, start_column=8, end_row=u_height_sheet, end_column=8)
                        sheet.merge_cells(start_row=height_device_in_sheet, start_column=9, end_row=u_height_sheet, end_column=9)
                        sheet.merge_cells(start_row=height_device_in_sheet, start_column=10, end_row=u_height_sheet, end_column=10)
                else:
                    # add item into sheet
                    empty_item = [rack,u_height_rack]
                    sheet.append(empty_item)
                    # print(f"add empty into rack {rack} and U{u_height_rack}")
                # variable counter
                
                u_height_rack-= 1
                u_height_sheet+= 1

        return workbook    
    except Exception as e:
        print("error export all device")
        return workbook

# Tìm device với rack name và position
def find_device_item(devices_list, rack_name, position):
    for device in devices_list:
        if device.rack == rack_name and device.position == position:
            return device
    return

# export tất cả các rack có trong hệ thống
def export_all_rack():
    # Create a new workbook and select the active worksheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Data Export"

    # create the headers
    headers = ['Rack', 'U/Rack','Tên Thiết bị', 'Chủng loại', 'Quản lý', 'Số HĐ', 'Model', 'SN', 'Thời gian lắp đặt', 'Ghi Chú']
    sheet.append(headers)
    
    try:
        # check_device = get_devices_child(654)
        racks_list = get_rack_have_device()
        devices_list = get_device()
        device_type_parent = get_device_type_parent()
        item_sheet_list = [] # danh sách item device đã xử lý dữ liệu
        # start loop for chuẩn bị dữ liệu export
        for device in devices_list:
            #Kiểm tra nếu thiết bị không nằm trong tủ rack hoặc không có position bỏ qua 
            if device.position is None:
                continue
            # get data custom feild
            device_owner = ""
            year_of_investment = ""
            contract_number = ""
            custom_fields = device.get_custom_fields_by_group()
            for key, value in custom_fields[''].items():
                if str(key) == 'Device owner' and value != None:
                    device_owner = value
                elif str(key) == 'Year of investment' and value != None:
                    year_of_investment = value
                elif str(key) == 'Contract number' and value != None:
                    contract_number = value

            # create new item export
            item_export = ExportExcel(
                id = device.id,
                rack = str(device.rack),
                device_role = str(device.role),
                device_type = str(device.device_type),
                device_name = str(device.name),
                position = int(device.position), # U/Rack
                serial_number = str(device.serial),
                device_description = str(device.description),
                owner_device = str(device_owner),
                year_of_investment = str(year_of_investment),
                contract_number = str(contract_number), 
                u_number = int(device.device_type.u_height),
            )
            item_sheet_list.append(item_export)
        # end loop for chuẩn bị dữ liệu export
        
        u_height_sheet = 2 # đếm dòng sheet dòng 1 là header nên bắt đầu từ 2 
        del_row_list = []
        # Start loop rack add raw into excel
        for rack in racks_list: 
            u_height_rack = 42 # mỗi 1 rack có 42 lần add item ( 1 item = 1 U)
            for i in range(42):
                # kiểm tra trong list item export
                device = find_device_item(item_sheet_list, rack.name, u_height_rack)
                if device:
                    # add item into sheet
                    item = [
                        rack.name,
                        u_height_rack,
                        device.device_name,
                        device.device_role,
                        device.owner_device,
                        device.contract_number,
                        device.device_type,
                        device.serial_number,
                        device.year_of_investment,
                        device.device_description,
                    ]
                    
                    # check height > 1 => merg cell 
                    if device.u_number > 1:
                        # kiểm tra xem device_type có phải là chassic không.
                        is_chassis = False
                        for device_type in device_type_parent:
                            if device_type.model == device.device_type:
                                is_chassis = True
                        
                        if is_chassis:
                            # Thêm các thiết bị con
                            device_child_list = get_devices_child(device.id)
                            
                            for device_child in device_child_list:
                                position_child = device_child.parent_bay.custom_field_data['position_bay']
                                # add item child
                                # get data custom feild
                                device_owner = ""
                                year_of_investment = ""
                                contract_number = ""
                                custom_fields = device_child.get_custom_fields_by_group()
                                for key, value in custom_fields[''].items():
                                    if str(key) == 'Device owner' and value != None:
                                        device_owner = value
                                    elif str(key) == 'Year of investment' and value != None:
                                        year_of_investment = value
                                    elif str(key) == 'Contract number' and value != None:
                                        contract_number = value
                                item_child = [
                                    rack.name,
                                    position_child,
                                    str(device_child.name),
                                    str(device_child.role),
                                    str(device_owner),
                                    str(contract_number), 
                                    str(device_child.device_type),
                                    str(device_child.serial),
                                    str(year_of_investment),
                                    str(device_child.description),
                                ]
                                sheet.append(item_child)
                                # copy giá trị của dòng hiện tại lên trên. 
                                height_device_in_sheet = u_height_sheet - device.u_number + 1 # chiều cao của thiết bị.
                                for col in range(2, 11):
                                    sheet.cell(row=height_device_in_sheet, column=col).value = sheet.cell(row=u_height_sheet, column=col).value
                                # Cộng 1 vào giá trị chiều cao của sheet
                                u_height_sheet+= 1
                            
                            
                            # Thay đổi position của cột và thêm item parent xuống cuối
                            end_position = u_height_rack + device.u_number - 1
                            item[1] = f"{u_height_rack}-{end_position}"
                            sheet.append(item)
                            
                            # merg cell của thiết bị cha
                            height_device_parrent_in_sheet = u_height_sheet - device.u_number + 1 # chiều cao của thiết bị.
                            # copy data từ row position sang ô đầu tiên
                            for col in range(2, 11):
                                sheet.cell(row=height_device_parrent_in_sheet, column=col).value = sheet.cell(row=u_height_sheet, column=col).value
                            
                            # Giảm biến đếm của excell về row có dữ liệu copy của chassis
                            # u_height_sheet = u_height_sheet - device.u_number + 1 
                            for row in range(height_device_parrent_in_sheet + 1,u_height_sheet +1):
                                del_row_list.append(row)
                            #merg cells.
                            # sheet.merge_cells(start_row=height_device_parrent_in_sheet, start_column=2, end_row=u_height_sheet, end_column=2)
                            # sheet.merge_cells(start_row=height_device_parrent_in_sheet, start_column=3, end_row=u_height_sheet, end_column=3)
                            # sheet.merge_cells(start_row=height_device_parrent_in_sheet, start_column=4, end_row=u_height_sheet, end_column=4)
                            # sheet.merge_cells(start_row=height_device_parrent_in_sheet, start_column=5, end_row=u_height_sheet, end_column=5)
                            # sheet.merge_cells(start_row=height_device_parrent_in_sheet, start_column=6, end_row=u_height_sheet, end_column=6)
                            # sheet.merge_cells(start_row=height_device_parrent_in_sheet, start_column=7, end_row=u_height_sheet, end_column=7)
                            # sheet.merge_cells(start_row=height_device_parrent_in_sheet, start_column=8, end_row=u_height_sheet, end_column=8)
                            # sheet.merge_cells(start_row=height_device_parrent_in_sheet, start_column=9, end_row=u_height_sheet, end_column=9)
                            # sheet.merge_cells(start_row=height_device_parrent_in_sheet, start_column=10, end_row=u_height_sheet, end_column=10)
                            
                            
                        else:
                            sheet.append(item)
                            # Từ cột thứ 3 đến cuối đều cần merg ô bằng chiều cao của thiết bị
                            height_device_in_sheet = u_height_sheet - device.u_number + 1 # chiều cao của thiết bị.
                            # copy data từ row position sang ô đầu tiên
                            for col in range(3, 11):
                                sheet.cell(row=height_device_in_sheet, column=col).value = sheet.cell(row=u_height_sheet, column=col).value
                            
                            #merg cells.
                            sheet.merge_cells(start_row=height_device_in_sheet, start_column=3, end_row=u_height_sheet, end_column=3)
                            sheet.merge_cells(start_row=height_device_in_sheet, start_column=4, end_row=u_height_sheet, end_column=4)
                            sheet.merge_cells(start_row=height_device_in_sheet, start_column=5, end_row=u_height_sheet, end_column=5)
                            sheet.merge_cells(start_row=height_device_in_sheet, start_column=6, end_row=u_height_sheet, end_column=6)
                            sheet.merge_cells(start_row=height_device_in_sheet, start_column=7, end_row=u_height_sheet, end_column=7)
                            sheet.merge_cells(start_row=height_device_in_sheet, start_column=8, end_row=u_height_sheet, end_column=8)
                            sheet.merge_cells(start_row=height_device_in_sheet, start_column=9, end_row=u_height_sheet, end_column=9)
                            sheet.merge_cells(start_row=height_device_in_sheet, start_column=10, end_row=u_height_sheet, end_column=10)
                            
                    else:
                        sheet.append(item)
                else:
                    # add item into sheet
                    empty_item = [rack.name,u_height_rack]
                    sheet.append(empty_item)
                
                u_height_rack-= 1
                u_height_sheet+= 1
        # End loop rack add raw into excel

        # Delete row in chassis
        del_row_list.sort(reverse=True)
        for number in del_row_list:
            sheet.delete_rows(number,1)
            
        return workbook 
    except Exception as e:
        print("error export all rack")
        return workbook
    
    
# export tất cả device 
def export_only_device():
    # Create a new workbook and select the active worksheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Data Export"

    # create the headers
    headers = ['Rack', 'Số U','Vị trí bắt đầu','Vị trí kế thúc', 'Tên Thiết bị', 'Chủng loại', 'Quản lý', 'Số HĐ', 'Model', 'SN', 'Thời gian lắp đặt', 'Ghi Chú']
    sheet.append(headers)

    try:

        # check form data: 1. export all table
        item_sheet_list = []
        # get all device
        devices_list = get_device()

        # start loop for
        for device in devices_list:
            # get data custom feild
            device_owner = ""
            year_of_investment = ""
            contract_number = ""
            custom_fields = device.get_custom_fields_by_group()
            for key, value in custom_fields[''].items():
                if str(key) == 'Device owner' and value != None:
                    device_owner = value
                elif str(key) == 'Year of investment' and value != None:
                    year_of_investment = value
                elif str(key) == 'Contract number' and value != None:
                    contract_number = value
            # Tính start U - end U
            end_u = int(device.position) + int(device.device_type.u_height) - 1
            # create new item export
            item_export = ExportExcel(
                rack = device.rack,
                device_role = device.role,
                device_type = device.device_type,
                device_name = device.name,
                position = int(device.position),
                serial_number = device.serial,
                device_description = device.description,
                owner_device = device_owner,
                year_of_investment = year_of_investment,
                contract_number = contract_number, 
                u_number = int(device.device_type.u_height),
                u_end = end_u,
            )
            
            # append data to item_sheet_list export
            item_sheet_list.append(item_export)

            # create item in sheet
            item_sheet = [
                str(item_export.rack), 
                str(item_export.u_number), 
                str(item_export.position), # U start
                str(item_export.u_end), # U end
                str(item_export.device_name),
                str(item_export.device_role),
                str(item_export.owner_device), 
                str(item_export.contract_number),
                str(item_export.device_type),
                str(item_export.serial_number),
                str(item_export.year_of_investment),
                str(item_export.device_description),
            ]
            sheet.append(item_sheet)
        # end loop for
        return workbook
    except Exception as e:
        print("return empty excel")
        return workbook